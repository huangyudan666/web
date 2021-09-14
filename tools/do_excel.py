from openpyxl import load_workbook
from tools.project_path import *
from tools.get_data import GetData
from tools.read_config import ReadConfig
from tools.do_regx import DoRegx

class DoExcel:
    def do_excel(self,file_name):
        wb=load_workbook(file_name)
        mode = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))
        tel=getattr(GetData,'tel')
        test_data=[]
        for key in mode:
            sheet=wb[key]
            if mode[key]=="all":
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data['case_id']=sheet.cell(i,1).value
                    row_data['url']=sheet.cell(i,2).value
                    if sheet.cell(i, 3).value.find('${tel}') != -1:  # 有找到这个${tel}
                        row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', str(tel))
                        tel =tel + 1 # 每次完成tel的调用后  就加1
                    else:
                        row_data['data']=DoRegx.do_regx(sheet.cell(i,3).value)
                    #sql语句的处理
                    if sheet.cell(i,4).value!=None:
                        row_data['check_sql']=DoRegx.do_regx(sheet.cell(i,4).value)
                    else:
                        row_data['check_sql']=None
                    row_data['title'] = sheet.cell(i, 5).value
                    row_data['http_method'] = sheet.cell(i, 6).value
                    row_data['excepted']=sheet.cell(i,7).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
                    self.update_tel(file_name,'init',tel)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['url'] = sheet.cell(case_id+1, 2).value
                    if sheet.cell(case_id+1,3).value.find('${tel}')!=-1:
                        row_data['data']=sheet.cell(case_id+1,3).value.replace('${tel}',str(tel))
                        tel = tel + 1
                    else:
                        row_data['data'] = DoRegx.do_regx(sheet.cell(case_id+1, 3).value)
                        # sql语句的处理
                    if sheet.cell(case_id+1, 4).value != None:
                            row_data['check_sql']= DoRegx.do_regx(sheet.cell(case_id+1, 4).value)
                    else:
                        row_data['check_sql']= None
                    row_data['title'] = sheet.cell(case_id + 1, 5).value
                    row_data['http_method'] = sheet.cell(case_id + 1, 6).value
                    row_data['excepted'] = sheet.cell(case_id+1, 7).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
                    self.update_tel(file_name, 'init', tel)
        return test_data

    def write_back(self,file_name,sheet_name,row,col,value):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(row,col).value=value
        wb.save(file_name)

    def update_tel(self,file_name,sheet_name,tel):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(file_name)

if __name__ == '__main__':
    print(DoExcel().do_excel(test_data_path))